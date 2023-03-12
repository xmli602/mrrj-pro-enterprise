#!/usr/bin/env python
#coding:utf-8

"""
@File    :   operationExcel.py
@Time    :   2023/01/28
@Author  :   zhuoyan
@Contact :   18108347985@163.com
@Desc    :   读取excel
"""

import openpyxl
from common.log import log  # 日志层

class OperationExcel(object):
    """
     读取文件方法
    """

    def ReadAllExcel(self, excelpath: str, sheetname: str):
        """
        获取所有用例
        :param excelpath: 文件路径 -->  Excel
        :param sheetname: 文件sheet名称
        :return: 返回表格所有数据列表
        """
        try:
            wb = openpyxl.load_workbook(excelpath)  # 加载工作簿
            sheet = wb[sheetname]  # 获得sheet对象
            testcase = []
            for row in range(2, sheet.max_row + 1):
                row_info = []
                [row_info.append(sheet.cell(row, col).value) for col in range(1, sheet.max_column + 1)]
                testcase.append(row_info)
            log.info('成功获取所有用例，共%d条用例' % (sheet.max_row - 1))
            return testcase
        except Exception as e:
            log.error("ReadAllExcel Error：{}".format(e))

    def ReadOneExcel(self, excelpath: str, sheetname: str, number: int):
        """
        获取某一条测试用例
        :param excelpath: 文件路径 -->  Excel
        :param sheetname: 文件sheet名称
        :param number: 索引 --> 从第二行开始为0的索引
        :return: 返回一行数据
        """
        try:
            wb = openpyxl.load_workbook(excelpath)  # 加载工作簿
            sheet = wb[sheetname]  # 获得sheet对象
            fist_list = []
            second_list = []
            for i in sheet[number + 1]:
                fist_list.append(i.value)
            second_list.append(fist_list)
            log.info(f'成功获取一条测试用例：{second_list}')
            return second_list
        except Exception as e:
            log.error("ReadOneExcel Error：{}".format(e))

    def RadeCellExcel(self, excelpath: str, sheetname: str, row: int, col: int):
        """
        读取具体某一行某一列数据（单元格）
        :param excelpath: 文件路径 -->  Excel
        :param sheetname: 文件sheet名称
        :param row: 行号 --> int
        :param col: 列 --> int
        :return: 单元格数据
        """
        try:
            wb = openpyxl.load_workbook(excelpath)  # 加载工作簿
            sheet = wb[sheetname]  # 获得sheet对象
            result = sheet.cell(row + 1, col + 1).value  # 输出具体单元格数据
            log.info(f'数据读取成功： {result}')
            return result
        except Exception as e:
            log.error("RadeCellExcel Error：{}".format(e))

    def WriteExcel(self, excelpath: str, sheetname: str, index: int, writevalue: str):
        """
        写入数据
        :param excelpath: 文件路径 -->  Excel
        :param sheetname: 文件sheet名称
        :param writevalue: 写入的数据
        :param index: 单元格位置
        :return: inster
        """
        try:
            wb = openpyxl.load_workbook(excelpath)  # 加载工作簿
            sheet = wb[sheetname]  # 获得sheet对象
            active = wb.active  # 写入文件时需要激活
            active[index] = str(writevalue)  # 通过单元格写入 例如：A1 A2 A3等
            wb.save(excelpath)  # 保存
            log.info('成功在%s表的%s页%s单元格写入数据内容为：%s' % (excelpath, sheetname, index, writevalue))
        except Exception as e:
            log.error("WriteExcel Error：{}".format(e))

    def RadeRowColExcel(self, excelpath: str, sheetname: str, row1: int, row2: int):
        """
        读取区间用例（例：取5-10行）
        :param excelpath: 文件路径 -->  Excel
        :param sheetname: 文件sheet名称
        :param row1: 行号 --> int
        :param row2: 行号 --> int
        :return: 返回区间数据
        """
        try:
            wb = openpyxl.load_workbook(excelpath)  # 加载工作簿
            sheet = wb[sheetname]  # 获得sheet对象
            all_list = []
            for rows in range(row1 + 1, sheet.max_row - row2):
                row_list = []
                for cols in range(1, sheet.max_column + 1):
                    row_list.append(sheet.cell(rows, cols).value)
                all_list.append(row_list)
            log.info('成功从获取第%d到%d用例,共%d条用例' % (
                row1 + 1, sheet.max_row - (row2 + 1), (sheet.max_row - row2) - row1 - 1))
            return all_list
        except Exception as e:
            log.error("RadeRowColExcel Error：{}".format(e))


# 实例化
GetExcel = OperationExcel()